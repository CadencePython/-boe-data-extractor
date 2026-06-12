"""
BOE (Bill of Entry) Data Extraction Pipeline
=============================================
Extracts all finance-relevant fields from Indian Customs BOE PDFs.
- Handles multiple PDFs in Input folder
- Handles multiple BOEs per PDF
- Handles multiple line items per BOE (produces one Excel row per item)

Folder structure:
  C:/Users/Admin/Downloads/BOE automation/
    Code/    <- this script
    Input/   <- drop all BOE PDFs here
    Output/  <- Excel output written here

Usage:  python boe_extractor.py
"""

import os, sys, re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── PATHS ─────────────────────────────────────────────────────────────────────
try:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
except NameError:
    _cwd = os.getcwd()
    BASE_DIR = os.path.dirname(_cwd) if os.path.basename(_cwd).lower() == "code" else _cwd

INPUT_DIR  = os.path.join(BASE_DIR, "Input")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output")
os.makedirs(INPUT_DIR,  exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

print(f"Base folder  : {BASE_DIR}")
print(f"Input folder : {INPUT_DIR}")
print(f"Output folder: {OUTPUT_DIR}")

# ── HELPERS ───────────────────────────────────────────────────────────────────

def clean(val):
    if val is None: return ""
    s = str(val).replace("\n", " ").strip()
    s = re.sub(r"^[A-Z]\s+", "", s).strip()   # strip single junk leading char
    return s

def find_all(table, keyword, partial=False):
    hits = []
    for r, row in enumerate(table):
        for c, cell in enumerate(row):
            if cell is None: continue
            s = str(cell).replace("\n", " ").strip()
            match = (keyword.lower() in s.lower()) if partial else (s == keyword)
            if match:
                hits.append((r, c))
    return hits

def find_cell(table, keyword, partial=False):
    hits = find_all(table, keyword, partial)
    return hits[0] if hits else (None, None)

def cell_val(table, row, col):
    if row is None or col is None: return ""
    if row < 0 or row >= len(table): return ""
    if col < 0 or col >= len(table[row]): return ""
    return clean(table[row][col])

def right(table, keyword, offset=1, partial=False):
    r, c = find_cell(table, keyword, partial)
    if r is None: return ""
    for o in range(offset, min(offset + 6, len(table[r]) - c)):
        v = cell_val(table, r, c + o)
        if v: return v
    return ""

def right_exact(table, keyword, offset, partial=False):
    r, c = find_cell(table, keyword, partial)
    return cell_val(table, r, None if c is None else c + offset)

def below(table, keyword, offset=1, partial=False):
    r, c = find_cell(table, keyword, partial)
    return cell_val(table, None if r is None else r + offset, c)

# ── PAGE TYPE DETECTION ───────────────────────────────────────────────────────

def page_has(table, text):
    for row in table:
        for cell in row:
            if cell and text in str(cell): return True
    return False

def is_part1(t): return page_has(t, "PART - I - BILL OF ENTRY SUMMARY")
def is_part2(t): return page_has(t, "PART - II - INVOICE")
def is_part3(t): return page_has(t, "PART - III - DUTIES")

# ── PART I ────────────────────────────────────────────────────────────────────
# Returns a single dict (header-level fields shared by all line items)

def extract_part1(table):
    d = {}

    # Header box scan
    d.update({"Port Code":"","BE No":"","BE Date":"","BE Type":"","BE Mode":""})
    for row in table[:8]:
        vals = {c: str(v).replace("\n"," ").strip() for c,v in enumerate(row) if v and str(v).strip()}
        if any(re.match(r"^INBLR", v) for v in vals.values()):
            for col_i, val in vals.items():
                if re.match(r"^INBLR\d*$", val):           d["Port Code"] = val
                elif re.match(r"^\d{6,8}$", val):           d["BE No"]     = val
                elif re.match(r"\d{2}/\d{2}/\d{4}", val):  d["BE Date"]   = val
                elif val in ("H","W","X","B"):               d["BE Type"]   = val
            break

    d["IEC/Br"]     = right_exact(table, "IEC/Br",     3)
    d["GSTIN/TYPE"] = right_exact(table, "GSTIN/TYPE", 3)
    d["CB CODE"]    = right_exact(table, "CB CODE",    3)
    d["G.WT (KGS)"] = right_exact(table, "G.WT (KGS)", 7)  # value is 7 cols right of label
    r_pkg, c_pkg    = find_cell(table, "PKG")
    d["PKG"]        = cell_val(table, r_pkg, c_pkg + 3) if r_pkg is not None else ""

    r_bs, _ = find_cell(table, "1.BE STATUS")
    d["BE Mode"] = ""
    if r_bs is not None and r_bs + 1 < len(table):
        for v in table[r_bs + 1]:
            if v and str(v).strip() in ("Air","Sea","Land","Road"):
                d["BE Mode"] = str(v).strip(); break

    d["Country of Origin"]      = right_exact(table, "13.COUNTRY OF ORIGIN",      5)
    d["Country of Consignment"] = right_exact(table, "14.COUNTRY OF CONSIGNMENT", 12)  # col 28
    d["Port of Loading"]        = right_exact(table, "15.PORT OF LOADING",          5)   # col 7
    d["Port of Shipment"]       = right_exact(table, "16.PORT OF SHIPMENT",         12)  # col 28

    r_imp, c_imp = find_cell(table, "1.IMPORTER NAME & ADDRESS")
    lines = []
    if r_imp is not None:
        for i in range(1, 7):
            tr = r_imp + i
            if tr >= len(table): break
            v = clean(table[tr][c_imp]) if c_imp < len(table[tr]) else ""
            if v and not any(k in v for k in ("2.CB NAME","3.AEO","4.UCR","AD CODE")):
                lines.append(v)
    d["Importer Name"]    = lines[0] if lines else ""
    d["Importer Address"] = ", ".join(lines[1:]) if len(lines) > 1 else ""

    r_cb, c_cb = find_cell(table, "2.CB NAME")
    d["CB Name"] = cell_val(table, r_cb, c_cb + 2) if r_cb is not None else ""

    duty_map = {
        "1.BCD":"BCD Total","2.ACD":"ACD Total","3.SWS":"SWS Total",
        "4.NCCD":"NCCD Total","5.ADD":"ADD Total","6.CVD":"CVD Total",
        "7.IGST":"IGST Total","8.G.CESS":"GCESS Total",
        "18.TOT.ASS VAL":"Total Assessable Value",
        "14.TOTAL DUTY":"Total Duty","15.INT":"Interest",
        "16.PNLTY":"Penalty","17.FINE":"Fine",
        "19.TOT. AMOUNT":"Total Amount Payable",
    }
    for label, field in duty_map.items():
        r_l, c_l = find_cell(table, label)
        d[field] = cell_val(table, r_l + 1 if r_l is not None else None, c_l)

    d["IGM No"]   = below(table, "1.IGM NO",   1)
    d["IGM Date"] = below(table, "2.IGM DATE", 1)
    d["INW Date"] = below(table, "3.INW DATE", 1)
    mawb = below(table, "6.MAWB NO", 1)
    d["MAWB No"] = re.sub(r"^[A-Z]\s+","", mawb).strip()
    hawb = below(table, "8.HAWB NO", 1)
    d["HAWB No"] = re.sub(r"\s+\d+$","", hawb).strip()

    r_ex, c_ex = find_cell(table, "EXCHANGE RATE")
    d["Exchange Rate"] = ""
    if r_ex is not None:
        for i in range(1, 5):
            v = cell_val(table, r_ex + i, c_ex)
            if re.search(r"\d.*=.*\d", v):
                d["Exchange Rate"] = re.sub(r"^[A-Z]\s+","", v).strip(); break

    d["Invoice No (Summary)"]  = below(table, "2.INVOICE NO", 1)
    d["Invoice Amt (Summary)"] = below(table, "3.INV. AMT",   1)
    d["Currency (Summary)"]    = below(table, "4.CUR",         1)

    d["OOC No"]   = right_exact(table, "OOC NO.",  4)
    d["OOC Date"] = right_exact(table, "OOC DATE", 4)

    return d

# ── PART II ───────────────────────────────────────────────────────────────────
# Returns a dict with:
#   - invoice-level fields (shared)
#   - "items": list of per-item dicts

def extract_part2(table):
    invoice = {}

    # Invoice header (shared across all items)
    invoice["Invoice No"]   = below(table, "2.INVOICE NO. & DT.",       1)
    invoice["Invoice Date"] = below(table, "2.INVOICE NO. & DT.",       2)
    invoice["PO No & Date"] = below(table, "3.PURCHASE ORDER NO & DT",  1)
    invoice["LC No & Date"] = below(table, "4.LC NO & DATE",             1)

    # Supplier
    r_sup, c_sup = find_cell(table, "3.SUPPLIER NAME & ADDRESS")
    lines = []
    if r_sup is not None:
        for i in range(1, 7):
            tr = r_sup + i
            if tr >= len(table): break
            v = cell_val(table, tr, c_sup)
            if v and not any(k in v for k in ("4.THIRD","5.AEO","6. AD")): lines.append(v)
    invoice["Supplier Name"]    = lines[0] if lines else ""
    invoice["Supplier Address"] = ", ".join(lines[1:]) if len(lines) > 1 else ""

    # Valuation (invoice level)
    invoice["Invoice Value"]    = below(table, "1.INV VALUE",        1)
    invoice["Freight"]          = below(table, "2.FREIGHT",          1)
    invoice["Insurance"]        = below(table, "3.INSURANCE",        1)
    invoice["Pay Terms"]        = below(table, "7.PAY TERMS",        1)
    invoice["Valuation Method"] = below(table, "8.VALUATION METHOD", 1)
    invoice["Assessable Value"] = below(table, "14.ASS. VALUE",      1)

    r_cur, c_cur = find_cell(table, "14.Cur")
    invoice["Currency"] = ""
    if r_cur is not None:
        for cc in range(c_cur + 1, min(c_cur + 6, len(table[r_cur]))):
            v = cell_val(table, r_cur, cc)
            if v in ("USD","EUR","GBP","INR","JPY","CNY","AUD","SGD","CHF","AED"):
                invoice["Currency"] = v; break

    r_tt, c_tt = find_cell(table, "15.Term")
    invoice["Trade Term"] = cell_val(table, r_tt, c_tt + 1) if r_tt is not None else ""

    # ── Item rows: find the E.ITEM DETAILS header row, then read each data row below ──
    # Header row has: 1.S NO. | 2.CTH | 3.DESCRIPTION | 4.UNIT PRICE | 5.QUANTITY | 6.UQC | 7.AMOUNT
    r_item_hdr, _ = find_cell(table, "1.S NO.")
    items = []
    if r_item_hdr is not None:
        # Map column positions from header row
        hdr_row = table[r_item_hdr]
        col_map = {}
        for c, v in enumerate(hdr_row):
            if v:
                col_map[str(v).replace("\n"," ").strip()] = c

        c_sno   = col_map.get("1.S NO.")
        c_cth   = col_map.get("2.CTH")
        c_desc  = col_map.get("A 3.DESCRIPTION") or col_map.get("3.DESCRIPTION")
        # fallback: find desc col by partial match
        if c_desc is None:
            for k, v in col_map.items():
                if "DESCRIPTION" in k: c_desc = v; break
        c_price = col_map.get("4.UNIT PRICE")
        c_qty   = col_map.get("5.QUANTITY")
        c_uqc   = col_map.get("6.UQC")
        c_amt   = col_map.get("7.AMOUNT")

        # Read rows below header until we hit a non-item row (glossary, blank section)
        r = r_item_hdr + 1
        while r < len(table):
            row = table[r]
            # Item rows start with a small integer in the S.NO column
            sno_val = cell_val(table, r, c_sno) if c_sno is not None else ""
            if re.match(r"^\d{1,3}$", sno_val):
                item = {
                    "Item S.No":          sno_val,
                    "CTH":                cell_val(table, r, c_cth)   if c_cth   is not None else "",
                    "Item Description":   cell_val(table, r, c_desc)  if c_desc  is not None else "",
                    "Unit Price":         cell_val(table, r, c_price) if c_price is not None else "",
                    "Quantity":           cell_val(table, r, c_qty)   if c_qty   is not None else "",
                    "UQC":                cell_val(table, r, c_uqc)   if c_uqc   is not None else "",
                    "Amount":             cell_val(table, r, c_amt)   if c_amt   is not None else "",
                }
                items.append(item)
            elif sno_val == "" and not any(cell_val(table, r, c_sno) for _ in [1]):
                # stop if we hit glossary or section break
                row_text = " ".join(str(v) for v in row if v)
                if "GLOSSARY" in row_text or "OUT OF" in row_text:
                    break
            r += 1

    # If no items found via header, fall back to single-item extraction
    if not items:
        items = [{
            "Item S.No":        "1",
            "CTH":              below(table, "2.CTH",         1),
            "Item Description": below(table, "3.DESCRIPTION", 1),
            "Unit Price":       below(table, "4.UNIT PRICE",  1),
            "Quantity":         below(table, "5.QUANTITY",    1),
            "UQC":              below(table, "6.UQC",         1),
            "Amount":           below(table, "7.AMOUNT",      1),
        }]

    return invoice, items

# ── PART III ──────────────────────────────────────────────────────────────────
# Returns list of per-item duty dicts (one per line item)

def extract_part3(table):
    """
    Finds all item blocks by scanning for "1.INVSNO" header rows.
    Each block: INVSNO header → item details rows → B.ITEM DUTY rows → C.OTHER DUTIES rows
    Returns list of dicts, one per item.
    """
    # Find all rows where col 2 = "1.INVSNO" (start of each item block)
    block_starts = []
    for r, row in enumerate(table):
        for c, cell in enumerate(row):
            if cell and str(cell).replace("\n"," ").strip() == "1.INVSNO":
                block_starts.append(r)
                break

    if not block_starts:
        return []

    # Add sentinel end
    block_starts.append(len(table))

    items = []
    for bi, block_start in enumerate(block_starts[:-1]):
        block_end = block_starts[bi + 1]
        # Slice the table to just this item's block
        block = table[block_start:block_end]

        d = {}

        # Item identification (row immediately after INVSNO header)
        if len(block) > 1:
            item_row = block[1]
            row_vals = {c: clean(v) for c, v in enumerate(item_row) if v and clean(v)}
            d["Item S.No (Duties)"]    = row_vals.get(3, "")   # ITEMSN col
            d["CTH (Duties)"]          = row_vals.get(4, "")
            d["CETH"]                  = row_vals.get(5, "")
            d["Item Desc (Duties)"]    = row_vals.get(6, "")

        # Assess value and total duty — find "29.ASSESS VALUE" row in this block
        r_av, c_av = find_cell(block, "29.ASSESS VALUE")
        if r_av is not None:
            d["Assess Value (Duties)"] = cell_val(block, r_av + 1, c_av)
        else:
            d["Assess Value (Duties)"] = ""

        r_td, c_td = find_cell(block, "30. TOTAL DUTY")
        if r_td is not None:
            d["Total Duty (Item)"] = cell_val(block, r_td + 1, c_td)
        else:
            d["Total Duty (Item)"] = ""

        # ── B.ITEM DUTY section ──
        r_bhdr, _ = find_cell(block, "1. BCD")
        cols_b = {}
        if r_bhdr is not None:
            for c, v in enumerate(block[r_bhdr]):
                if v: cols_b[str(v).replace("\n"," ").strip()] = c

        # Find Notn No. rows in this block
        notn_hits = find_all(block, "Notn No.")
        r_notn_b = notn_hits[0][0] if notn_hits else None
        r_notn_c = notn_hits[1][0] if len(notn_hits) > 1 else None

        def get_duty(r_notn, col_key, cols_dict):
            col = cols_dict.get(col_key)
            if col is None or r_notn is None: return "", "", "", ""
            notn_no  = cell_val(block, r_notn,   col)
            notn_sno = cell_val(block, r_notn+1, col)
            # Rate row can be at +2 or +3 (blank row position varies by BOE)
            # Scan offsets +2 to +4 for a numeric Rate value
            rate = ""
            rate_offset = None
            for off in [2, 3, 4]:
                v = cell_val(block, r_notn + off, col)
                if re.search(r"\d", v):
                    rate = v; rate_offset = off; break
            # Amount is always 2 rows after Rate (skip one row between Rate and Amount)
            amt = ""
            if rate_offset is not None:
                for off in [rate_offset + 1, rate_offset + 2]:
                    v = cell_val(block, r_notn + off, col)
                    if re.search(r"\d", v):
                        amt = v; break
            return notn_no, notn_sno, rate, amt

        d["BCD Notn No"],  d["BCD Notn SNo"],  d["BCD Rate"],  d["BCD Amount"]  = get_duty(r_notn_b, "1. BCD",   cols_b)
        _,                 _,                   d["SWS Rate"],  d["SWS Amount"]  = get_duty(r_notn_b, "3.SWS",   cols_b)

        igst_key  = next((k for k in cols_b if "5.IGST"  in k), None)
        gcess_key = next((k for k in cols_b if "6.G. CESS" in k), None)
        d["IGST Notn No"],  d["IGST Notn SNo"],  d["IGST Rate"],  d["IGST Amount"]  = get_duty(r_notn_b, igst_key,  cols_b) if igst_key  else ("","","","")
        d["GCESS Notn No"], d["GCESS Notn SNo"], d["GCESS Rate"], d["GCESS Amount"] = get_duty(r_notn_b, gcess_key, cols_b) if gcess_key else ("","","","")

        # ── C.OTHER DUTIES ──
        r_chdr, _ = find_cell(block, "5.CAIDC")
        cols_c = {}
        if r_chdr is not None:
            for c, v in enumerate(block[r_chdr]):
                if v: cols_c[str(v).replace("\n"," ").strip()] = c
        caidc_key = next((k for k in cols_c if "5.CAIDC" in k), None)
        nn, ns, rate, amt = get_duty(r_notn_c, caidc_key, cols_c) if caidc_key else ("","","","")
        # In some BOEs CAIDC actual amount is in Duty Fg row — check if amt looks wrong
        if caidc_key and r_notn_c is not None and cols_c.get(caidc_key) is not None:
            c_ca = cols_c[caidc_key]
            # Duty Fg row is at r_notn_c + 5 — grab it if Amount is 0 or empty
            duty_fg = cell_val(block, r_notn_c + 4, c_ca)  # Duty Fg row = Notn No + 4
            if (not amt or amt in ("0","0.0")) and re.search(r"[1-9]", duty_fg):
                amt = duty_fg
        d["CAIDC Notn No"], d["CAIDC Notn SNo"], d["CAIDC Rate"], d["CAIDC Amount"] = nn, ns, rate, amt

        items.append(d)

    return items

# ── MULTI-BOE / MULTI-ITEM READER ─────────────────────────────────────────────
# Returns list of records — one record per LINE ITEM across all BOEs in the PDF

def extract_boes_from_pdf(pdf_path):
    all_records = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        print(f"  {total} pages in PDF")

        # Find pages starting each BOE (Part I pages)
        starts = [i for i, pg in enumerate(pdf.pages)
                  if pg.extract_tables() and is_part1(pg.extract_tables()[0])]

        if not starts:
            print("  WARNING: No Part I page found — skipping")
            return []

        print(f"  {len(starts)} BOE(s) detected (Part I at page(s): {[s+1 for s in starts]})")
        starts.append(total)  # sentinel

        for boe_idx in range(len(starts) - 1):
            start, end = starts[boe_idx], starts[boe_idx + 1]
            print(f"  BOE {boe_idx+1}: pages {start+1}–{end}")

            boe_header = {"Source PDF": os.path.basename(pdf_path), "BOE Index": boe_idx + 1}
            p1_data = {}
            p2_invoice = {}
            p2_items   = []
            p3_items   = []

            for pg in pdf.pages[start:end]:
                tbls = pg.extract_tables()
                if not tbls: continue
                t = tbls[0]
                if   is_part1(t) and not p1_data:
                    p1_data = extract_part1(t)
                elif is_part2(t) and not p2_invoice:
                    p2_invoice, p2_items = extract_part2(t)
                elif is_part3(t) and not p3_items:
                    p3_items = extract_part3(t)

            # Determine number of line items (max across Part II and III)
            n_items = max(len(p2_items), len(p3_items), 1)
            print(f"    {n_items} line item(s) found")

            for item_idx in range(n_items):
                rec = {}
                rec.update(boe_header)
                rec["Item Index"] = item_idx + 1
                rec.update(p1_data)
                rec.update(p2_invoice)
                # Merge item-level data from Part II and III if available
                if item_idx < len(p2_items):
                    rec.update(p2_items[item_idx])
                if item_idx < len(p3_items):
                    rec.update(p3_items[item_idx])
                all_records.append(rec)

    return all_records

# ── EXCEL OUTPUT ──────────────────────────────────────────────────────────────

COLUMN_ORDER = [
    # Identification
    "Source PDF","BOE Index","Item Index",
    # BE Header
    "Port Code","BE No","BE Date","BE Type","BE Mode",
    "IEC/Br","GSTIN/TYPE","CB CODE","G.WT (KGS)","PKG",
    # Parties
    "Importer Name","Importer Address","CB Name",
    "Supplier Name","Supplier Address",
    # Origin & Routing
    "Country of Origin","Country of Consignment",
    "Port of Loading","Port of Shipment",
    # Manifest
    "IGM No","IGM Date","INW Date","MAWB No","HAWB No",
    # Invoice (shared)
    "Invoice No","Invoice Date","PO No & Date","LC No & Date",
    "Invoice Value","Currency","Freight","Insurance",
    "Pay Terms","Trade Term","Valuation Method","Assessable Value",
    # Item-level (Part II)
    "Item S.No","CTH","Item Description","Unit Price","Quantity","UQC","Amount",
    # Valuation
    "Exchange Rate",
    # Duty Summary (Part I totals)
    "Total Assessable Value","BCD Total","ACD Total","SWS Total",
    "NCCD Total","ADD Total","CVD Total","IGST Total","GCESS Total",
    "Total Duty","Interest","Penalty","Fine","Total Amount Payable",
    # Item-level duties (Part III)
    "Item S.No (Duties)","CTH (Duties)","CETH","Item Desc (Duties)",
    "Assess Value (Duties)","Total Duty (Item)",
    "BCD Notn No","BCD Notn SNo","BCD Rate","BCD Amount",
    "SWS Rate","SWS Amount",
    "IGST Notn No","IGST Notn SNo","IGST Rate","IGST Amount",
    "GCESS Notn No","GCESS Notn SNo","GCESS Rate","GCESS Amount",
    "CAIDC Notn No","CAIDC Notn SNo","CAIDC Rate","CAIDC Amount",
    # OOC
    "OOC No","OOC Date",
    # Invoice summary
    "Invoice No (Summary)","Invoice Amt (Summary)","Currency (Summary)",
]

GROUPS = [
    ("Identification",3),("BE Header",10),("Parties",5),
    ("Origin & Routing",4),("Manifest",5),("Invoice (Shared)",12),
    ("Item Details",7),("Valuation",1),("Duty Summary (Part I)",13),
    ("Item Duties (Part III)",22),("OOC",2),("Invoice Summary",3),
]
GROUP_COLORS = [
    "4472C4","70AD47","ED7D31","FFC000","5B9BD5",
    "843C0C","C55A11","375623","C00000","7030A0",
    "203864","1F497D",
]

def write_excel(records, path):
    all_keys = {k for r in records for k in r}
    extra    = [k for k in sorted(all_keys) if k not in COLUMN_ORDER]
    cols     = COLUMN_ORDER + extra

    wb = Workbook(); ws = wb.active; ws.title = "BOE Data"
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: group headers
    col_ptr = 1
    for (grp, width), color in zip(GROUPS, GROUP_COLORS):
        end_col = min(col_ptr + width - 1, len(cols))
        ws.merge_cells(start_row=1,start_column=col_ptr,end_row=1,end_column=end_col)
        cell = ws.cell(row=1, column=col_ptr, value=grp)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        col_ptr        = end_col + 1

    # Row 2: field headers
    for ci, name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=name)
        cell.fill      = PatternFill("solid", fgColor="D9E1F2")
        cell.font      = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # Data rows
    for ri, rec in enumerate(records, 3):
        for ci, name in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=rec.get(name, ""))
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")

    for ci, name in enumerate(cols, 1):
        max_len = max(len(name), max((len(str(r.get(name,""))) for r in records), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 42)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "D3"
    wb.save(path)
    print(f"  Saved → {path}")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    pdfs = sorted(f for f in os.listdir(INPUT_DIR) if f.lower().endswith(".pdf"))
    if not pdfs:
        print(f"No PDFs found in {INPUT_DIR}"); sys.exit(1)

    print(f"\n{'─'*60}")
    print(f"BOE Extractor  |  {len(pdfs)} PDF(s) found")
    print(f"{'─'*60}\n")

    all_records = []
    for pdf_file in pdfs:
        print(f"▶ {pdf_file}")
        try:
            recs = extract_boes_from_pdf(os.path.join(INPUT_DIR, pdf_file))
            all_records.extend(recs)
            print(f"  ✓ {len(recs)} row(s) extracted\n")
        except Exception as e:
            print(f"  ✗ ERROR: {e}\n")
            import traceback; traceback.print_exc()

    if not all_records:
        print("No records extracted."); sys.exit(1)

    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(OUTPUT_DIR, f"BOE_Extracted_{ts}.xlsx")
    print(f"Writing {len(all_records)} row(s) to Excel...")
    write_excel(all_records, out)
    print(f"\n{'─'*60}")
    print(f"✅  {len(all_records)} row(s) from {len(pdfs)} PDF(s)")
    print(f"   Output → {out}")
    print(f"{'─'*60}\n")

if __name__ == "__main__":
    main()
